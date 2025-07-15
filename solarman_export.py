import requests
import io
import openpyxl
import os
import logging
from datetime import datetime, timedelta
from telegram_message_sender import send_telegram_message
from telegram_message_formatter import make_report, make_report_month

# Logging configuration
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('logs/solarman_bot.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Create logs directory
os.makedirs('logs', exist_ok=True)

url = "https://home.solarmanpv.com/maintain-s/station/report/export/63607492"

# Get token from environment variables or use default value
_TOKEN: str = os.getenv("SOLARMAN_TOKEN", "eyJhbGciOiJSUzI1NiIsInR5cCI6IkpXVCJ9.eyJ1c2VyX25hbWUiOiIwX3ZpdGFsaXkuc2hhcmFldnNraXlAZ21haWwuY29tXzIiLCJtb2RpZnlfcGFzc3dvcmQiOjEsInNjb3BlIjpbImFsbCJdLCJkZXRhaWwiOnsib3JnYW5pemF0aW9uSWQiOjAsInRvcEdyb3VwSWQiOm51bGwsImdyb3VwSWQiOm51bGwsInJvbGVJZCI6LTEsInVzZXJJZCI6MTQ5MDg3NDksInZlcnNpb24iOjEwMDAsImlkZW50aWZpZXIiOiJ2aXRhbGl5LnNoYXJhZXZza2l5QGdtYWlsLmNvbSIsImlkZW50aXR5VHlwZSI6MiwibWRjIjoiRk9SRUlHTl8xIiwiYXBwSWQiOm51bGx9LCJleHAiOjE3NTUzMzg1NDYsIm1kYyI6IkZPUkVJR05fMSIsImF1dGhvcml0aWVzIjpbImFsbCJdLCJqdGkiOiI5ZTZjMjJhMi01MDVmLTQyZmMtYjE3ZS0wZTg0Y2IwMmU1M2IiLCJjbGllbnRfaWQiOiJ0ZXN0In0.brSkfdiXXope_M8JPoEF7ps6hO_675ws0RrkqBryFN5Zjq9Ntr8fjJow-iAaz_5-oyjOJF_BSAJ6kL-GpgQJR0_d7Dgs1Ul29gnn2eqNxx0z6y8xdwHESlx7VBQKzqoUR8TpfQy9y1XmTBl-lP9oy99Asy3gIWtsNIB-tJrLQqElwS0IKVp4lQfUtu7rFV70SAcLiz3Uls2CqYPo4UyUN1fftSv7yYuKcfl4oVzucYq-Q48qcS8LwcAhw-DH4P7jtkxxiqxyEXv48EJJPoABtxkhZtpByJNWi1YD_iGB2umSJC8xqcu_xEylKAnkrjfEiNL5ojbQ1Xo4iccjNwMF4Q")

headers = {
    "Authorization": f"Bearer {_TOKEN}",
    "Cookie": (
        "_ga=GA1.2.1999936839.1750146529; "
        "_ga_PD9MYSPQK5=GS2.2.s1750154483$o3$g1$t1750154498$j45$l0$h0; "
        "_gid=GA1.2.725920252.1750146529; "
        "accountFirstUse=eMail; "
        "affixPath=/business/maintain/plant; "
        "HMACCOUNT=C39DF37E2ABFFB27; "
        "language=uk; "
    ),
    "Content-Type": "application/json;charset=UTF-8",
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/137.0.0.0 Safari/537.36"
}

def get_solar_data():
    """Get data from Solarman API"""
    yesterday = datetime.now() - timedelta(days=1)
    date_str = yesterday.strftime('%Y-%m-%d')
    display_date_str = yesterday.strftime('%d.%m.%Y')

    payload_day = {
        "timeType":"2",
        "startTime":date_str,
        "endTime":date_str,
        "reportFields": [
            {
                "name":"GENERATE_VALUE",
                "storageName":"GENERATE_VALUE"
            },
            {
                "name":"USE_VALUE",
                "storageName":"USE_VALUE"
            },
            {
                "name":"BUY_VALUE",
                "storageName":"BUY_VALUE"
            },
            {
                "name":"CHARGE_VALUE",
                "storageName":"CHARGE_VALUE"
            },
            {
                "name":"DISCHARGE_VALUE",
                "storageName":"DISCHARGE_VALUE"
            },
            {
                "name":"PRE_INCOME",
                "storageName":"PRE_INCOME"
            }
        ]
    }

    try:
        logger.info(f"Requesting data for {display_date_str}")
        response = requests.post(url, headers=headers, json=payload_day, timeout=30)
        
        if response.ok:
            excel_bytes = io.BytesIO(response.content)
            wb = openpyxl.load_workbook(excel_bytes)
            ws = wb.active

            # Assume first row as headers, second row as values
            data = {}
            keys = [str(cell.value) if cell.value is not None else "" for cell in ws[1]]   # 1st row
            values = [cell.value for cell in ws[2]] # 2nd row
            data = dict(zip(keys, values))

            buy_value = data.get("BUY_VALUE", 0)
            use_value = data.get("USE_VALUE", 0)
            
            try:
                buy = float(buy_value) if buy_value is not None else 0.0
                use = float(use_value) if use_value is not None else 0.0
            except (ValueError, TypeError):
                buy = 0.0
                use = 0.0

            if use == 0:
                data["GENERATION_RATIO"] = "N/A"
            else:
                data["GENERATION_RATIO"] = round((buy / use) * 100)

            report = make_report(data, display_date_str)
            logger.info("Data successfully retrieved and processed")
            return report
        else:
            logger.error(f"API Error: {response.status_code} - {response.text}")
            return None
            
    except Exception as e:
        logger.error(f"Error retrieving data: {str(e)}")
        return None

def main():
    """Main function"""
    logger.info("Starting Solarman Bot")
    
    report = get_solar_data()
    
    if report:
        try:
            #send_telegram_message(report)
            logger.info("Report successfully sent to Telegram: " + report)
        except Exception as e:
            logger.error(f"Error sending to Telegram: {str(e)}")
    else:
        logger.error("Failed to get data for report")

if __name__ == "__main__":
    main()
